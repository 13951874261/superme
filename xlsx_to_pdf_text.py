from __future__ import annotations

import sys
import textwrap
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas

pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))

PAGE = landscape(A4)
PAGE_W, PAGE_H = PAGE
MARGIN = 12 * mm
TITLE_FONT = 'STSong-Light'
BODY_FONT = 'STSong-Light'
TITLE_SIZE = 12
BODY_SIZE = 7.5
LINE_H = BODY_SIZE * 1.45


def cell_text(value):
    if value is None:
        return ''
    return str(value).replace('\r', ' ').replace('\n', ' ').strip()


def row_text(row):
    vals = [cell_text(v) for v in row]
    while vals and vals[-1] == '':
        vals.pop()
    return ' | '.join(vals)


def draw_header(c, title, page_no):
    c.setFont(TITLE_FONT, TITLE_SIZE)
    c.drawString(MARGIN, PAGE_H - MARGIN, title)
    c.setFont(BODY_FONT, 7)
    c.drawRightString(PAGE_W - MARGIN, MARGIN / 2, f'Page {page_no}')
    return PAGE_H - MARGIN - 9 * mm


def new_page(c, title, page_no):
    c.showPage()
    return draw_header(c, title, page_no)


def wrap_text(text, chars):
    if not text:
        return ['']
    return textwrap.wrap(text, width=chars, break_long_words=True, break_on_hyphens=False) or ['']


def convert(src: Path, out: Path):
    wb = load_workbook(src, read_only=True, data_only=True)
    out.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(out), pagesize=PAGE)
    page_no = 1
    first_sheet = True

    usable_w = PAGE_W - 2 * MARGIN
    # Chinese glyphs are approximately square at this font size; keep line width conservative.
    max_chars = max(70, int(usable_w / (BODY_SIZE * 0.72)))

    try:
        for ws in wb.worksheets:
            title = f'{src.stem} - {ws.title}'
            if not first_sheet:
                c.showPage()
                page_no += 1
            first_sheet = False
            y = draw_header(c, title, page_no)
            c.setFont(BODY_FONT, BODY_SIZE)

            for row in ws.iter_rows(values_only=True):
                text = row_text(row)
                if not text:
                    continue
                lines = wrap_text(text, max_chars)
                needed = len(lines) * LINE_H + 2
                if y - needed < MARGIN:
                    page_no += 1
                    y = new_page(c, title, page_no)
                    c.setFont(BODY_FONT, BODY_SIZE)
                for line in lines:
                    c.drawString(MARGIN, y, line)
                    y -= LINE_H
                y -= 2
        c.save()
    finally:
        wb.close()


def main():
    if len(sys.argv) != 3:
        print('usage: xlsx_to_pdf_text.py <src.xlsx> <out.pdf>')
        return 2
    src = Path(sys.argv[1])
    out = Path(sys.argv[2])
    convert(src, out)
    print(f'OK {out} {out.stat().st_size}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
