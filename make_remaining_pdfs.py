from __future__ import annotations

import sys
import textwrap
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

FONT = 'Courier'
FONT_SIZE = 7
TITLE_FONT_SIZE = 12
MARGIN = 12 * mm
LINE_GAP = 1.35
PAGE_SIZE = landscape(A4)
PAGE_W, PAGE_H = PAGE_SIZE


def wrap_row(values: list[object], max_chars: int) -> list[str]:
    parts = []
    for v in values:
        if v is None:
            parts.append('')
        else:
            s = str(v).replace('\r', ' ').replace('\n', ' ').strip()
            parts.append(s)
    line = ' | '.join(parts).rstrip()
    if not line:
        return ['']
    return textwrap.wrap(line, width=max_chars, break_long_words=True, break_on_hyphens=False) or ['']


def write_sheet(c: canvas.Canvas, workbook_name: str, sheet_name: str, ws) -> None:
    title = f'{workbook_name} / {sheet_name}'
    c.setFont('Helvetica-Bold', TITLE_FONT_SIZE)
    y = PAGE_H - MARGIN
    c.drawString(MARGIN, y, title)
    y -= 10 * mm

    usable_width = PAGE_W - 2 * MARGIN
    char_w = stringWidth('W', FONT, FONT_SIZE)
    max_chars = max(30, int(usable_width / char_w))

    c.setFont(FONT, FONT_SIZE)

    for row in ws.iter_rows(values_only=True):
        wrapped = wrap_row(list(row), max_chars)
        for text in wrapped:
            if y < MARGIN + 12:
                c.showPage()
                c.setFont('Helvetica-Bold', TITLE_FONT_SIZE)
                y = PAGE_H - MARGIN
                c.drawString(MARGIN, y, title)
                y -= 10 * mm
                c.setFont(FONT, FONT_SIZE)
            c.drawString(MARGIN, y, text[:max_chars])
            y -= FONT_SIZE * LINE_GAP


def make_pdf(src: Path) -> Path:
    root = src.parents[2]
    out_dir = root / 'pdf'
    out_dir.mkdir(exist_ok=True)
    out_name = f'{src.parent.parent.name}_{src.stem}.pdf'
    out_path = out_dir / out_name

    wb = load_workbook(src, read_only=True, data_only=True)
    try:
        c = canvas.Canvas(str(out_path), pagesize=PAGE_SIZE)
        for ws in wb.worksheets:
            write_sheet(c, src.name, ws.title, ws)
            c.showPage()
        c.save()
    finally:
        wb.close()

    return out_path


def main() -> int:
    if len(sys.argv) < 2:
        print('usage: make_remaining_pdfs.py <missing_sources.txt>')
        return 1

    missing_list = Path(sys.argv[1])
    paths = [Path(line.strip()) for line in missing_list.read_text(encoding='utf-8').splitlines() if line.strip()]
    for src in paths:
        out = make_pdf(src)
        print(f'OK {out}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
